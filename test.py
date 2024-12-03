import pdfplumber
import pandas as pd
import re
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import glob
import numpy as np


def get_college_name(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if "BASIC INFORMATION" in text:
                    tables = page.extract_tables()
                    for table in tables:
                        # Convert table cells to strings and check if we found the right table
                        str_table = [[str(cell).strip() if cell is not None else "" for cell in row] for row in table]
                        for i, row in enumerate(str_table):
                            # Look for the row with college name (typically second row)
                            if i > 0 and len(row) > 1:  # Skip header row
                                college_name = row[1].strip()
                                if college_name and college_name.lower() != "name of the college":
                                    return college_name
    except Exception as e:
        print(f"Error extracting college name from {pdf_path}: {str(e)}")
    return None

def get_pdf_files(folder_path):

    pdf_pattern = os.path.join(folder_path, '**', '*.[pP][dD][fF]')
    pdf_files = glob.glob(pdf_pattern, recursive=True)
    
    if not pdf_files:
        print(f"No PDF files found in: {folder_path}")
        return []
    
    print(f"Found {len(pdf_files)} PDF files")
    return pdf_files

def extract_table(pdf_path, table_index=0):
    
    files_with_heading_no_table = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Pattern to match the specific heading 
            heading_pattern = r'((?:\d+\.)+\d+)\s*Number\s+of\s+complaints/grievances\s+about\s+evaluation\s+year\s+wise\s+during\s+last\s+five\s+years'
            
            # Track if heading is found and table is successfully extracted
            heading_found = False
            table_extracted = False
            
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                
                # Check if the heading is on this page
                match = re.search(heading_pattern, text, re.IGNORECASE)
                
                if match:
                    heading_found = True
                    section_number = match.group(1)
                    
                    # Extract tables on this page
                    tables = page.extract_tables()
                    
                    # Check if we have enough tables to extract the specified index
                    if tables and len(tables) > table_index:
                        table = tables[table_index]
                        
                        # Check if the table has the expected structure
                        if len(table) >= 2 and len(table[0]) == 5:
                            first_row = table[0]
                            
                            # Verify the first row contains year-like patterns
                            if all(re.match(r'\d{4}-\d{2}', str(cell)) for cell in first_row):
                                df = pd.DataFrame(table[1:], columns=table[0])
                                df = df.apply(pd.to_numeric, errors='ignore')
                                
                                print(f"Found table in {pdf_path} on page {page_num + 1}, table index {table_index}")
                                table_extracted = True
                                return df, section_number, files_with_heading_no_table
                    
                    # If we reached here, heading was found but table couldn't be extracted
                    if not table_extracted:
                        files_with_heading_no_table.append((pdf_path, section_number))
                    
                    break 
    
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")
    
    return None, None, files_with_heading_no_table

def save_to_excel(dataframes, output_path):

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        workbook = writer.book
        
        # Create main data sheet
        worksheet = workbook.create_sheet("College Data", 0)
        
        # Starting column for first college
        current_col = 1
        
        # Default years to use when no data is available
        default_years = ['x-x', 'x-x', 'x-x', 'x-x' , 'x-x']
        
        # Process each college's data
        for pdf_path, (df, section_number) in dataframes.items():
            # Get college name
            college_name = get_college_name(pdf_path) or os.path.splitext(os.path.basename(pdf_path))[0]
            
            # Write college name
            cell = worksheet.cell(row=1, column=current_col)
            cell.value = college_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
            # Merge cells for college name across its years
            worksheet.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + 4  # Merge 5 cells for years
            )
            
            # Determine years and data to use
            if df is not None and len(df) > 0:
                # Use actual data from the dataframe
                years = df.columns
                data_row = df.iloc[0]
            else:
                # Use default years and 'x' for data
                years = default_years
                data_row = ['x'] * len(default_years)
            
            # Write years (row 2)
            for i, year in enumerate(years):
                cell = worksheet.cell(row=2, column=current_col + i)
                cell.value = year
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Write data (row 3)
            for i, value in enumerate(data_row):
                cell = worksheet.cell(row=3, column=current_col + i)
                cell.value = value
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for i in range(5):  # 5 columns for each college
                col_letter = get_column_letter(current_col + i)
                worksheet.column_dimensions[col_letter].width = 15
            
            # Move to next college's columns
            current_col += 5
        
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])


def process_folder(input_folder, output_dir, table_index=0):
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Get all PDF files from the folder
    pdf_files = get_pdf_files(input_folder)
    
    if not pdf_files:
        return
    
    # Process all PDFs
    results = {}
    files_with_heading_no_table = []
    for pdf_path in pdf_files:
        try:
            print(f"Processing: {pdf_path}")
            df, section_number, no_table_files = extract_table(pdf_path, table_index)
            results[pdf_path] = (df, section_number)
            files_with_heading_no_table.extend(no_table_files)
            
            if df is not None:
                print(f"✓ Successfully extracted table from: {os.path.basename(pdf_path)}")
            else:
                print(f"✗ No matching table found in: {os.path.basename(pdf_path)}")
        except Exception as e:
            print(f"Error processing {pdf_path}: {str(e)}")
            results[pdf_path] = (None, None)
    
    # Timestamp for output files
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Save to Excel (ALL files, now including those without data)
    excel_path = os.path.join(output_dir, f'listed_{timestamp}.xlsx')
    save_to_excel(results, excel_path)
    print(f"Excel output saved to: {excel_path}")
    
    # Print processing summary
    print(f"\nProcessing Summary:")
    print(f"Total PDFs processed: {len(pdf_files)}")
    successful_extractions = sum(1 for _, (df, _) in results.items() if df is not None)
    print(f"Successfully extracted tables: {successful_extractions}")
    print(f"Failed extractions: {len(pdf_files) - successful_extractions}")
    
    # Print files with heading but no extractable table
    if files_with_heading_no_table:
        print("\nFiles with heading but no extractable table:")
        for file, section_number in files_with_heading_no_table:
            print(f"- {os.path.basename(file)} (Section: {section_number})")
    else:
        print("\nNo PDF files were found to process")


if __name__ == "__main__":
   
    input_folder = "ALL"  # Folder containing PDFs
    output_dir = "heading_complaints"   # Where to save the results
    process_folder(input_folder, output_dir, table_index=2)
